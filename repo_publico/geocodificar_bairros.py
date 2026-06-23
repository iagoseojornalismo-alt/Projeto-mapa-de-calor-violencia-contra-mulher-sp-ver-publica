#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Geocodifica os BAIRROS das ocorrências sem latitude/longitude na planilha,
posicionando-as corretamente no mapa em nível de bairro.

ESTRATÉGIA EM 3 CAMADAS (em ordem de prioridade):
  1. Coordenadas reais da planilha
     → usa diretamente quando lat/lng presentes e válidos na linha da ocorrência.

  2. Match por variante de grafia (opt-in via FUZZY_PREFIX)
     → se FUZZY_PREFIX > 0, associa bairros sem coord a um bairro geolocado da mesma
        cidade cujo nome normalizado comece com os mesmos FUZZY_PREFIX caracteres.
       Ex.: "ALFREDO BENEDETI" → "ALFREDO BENEDETTI" (typo leve).
       ATENÇÃO: valores baixos (< 12) geram falsos positivos. Use 0 para desativar.

  3. Geocodificação online via Nominatim (OpenStreetMap)
     → para tudo que não casou nas camadas anteriores. Salva em cache local para
        reaproveitamento — você pode interromper (Ctrl+C) e retomar.

COMO USAR (computador com acesso à internet):
    pip install openpyxl requests
    python geocodificar_bairros.py

Coloque NA MESMA PASTA:
    - a planilha .xlsx da SSP  (ajuste XLSX abaixo se necessário)
    - o arquivo mapa_violencia_mulher_sp.html

Saídas geradas/atualizadas:
    - geocode_cache.json              → cache reutilizável
    - mapa_violencia_mulher_sp.html   → mapa com bairros reposicionados

NOTAS:
  * Nominatim público: ~1 req/s. ~19 k bairros residuais ≈ 5–6 h na 1ª execução.
    O cache torna as seguintes quase instantâneas.
  * Para acelerar: use instância própria do Nominatim/Photon (troque GEOCODER_URL).
  * Bairros não encontrados ficam no centróide do município (comportamento anterior).
"""

import os, re, json, time, unicodedata
from collections import defaultdict

# ──────────────────────────── CONFIG ────────────────────────────
XLSX     = "FALASP_2026052710502690_-_violência_doméstica_com_bairros.xlsx"
HTML_IN  = "mapa_violencia_mulher_sp.html"
HTML_OUT = "mapa_violencia_mulher_sp.html"
CACHE    = "geocode_cache.json"

GEOCODER_URL = "https://nominatim.openstreetmap.org/search"
# Política do Nominatim: User-Agent identificável. Troque o e-mail abaixo!
USER_AGENT   = "mapa-violencia-mulher-sp/1.0 (contato: seu-email@exemplo.com)"
SLEEP_SEC    = 1.1       # ~1 req/s — respeite o rate-limit do Nominatim público
SP_BOUNDS    = (-25.6, -19.5, -53.6, -44.0)   # lat_min, lat_max, lng_min, lng_max
SAVE_EVERY   = 25        # grava o cache a cada N novas geocodificações online

# Camada 2 — match por prefixo normalizado.
# 0 = desativado (mais seguro); 12+ = ativado (pode ter falsos positivos).
FUZZY_PREFIX = 0

# Filtra cidades para testes rápidos (lista vazia = todas).
SOMENTE_CIDADES = []   # ex.: ["S.PAULO", "CAMPINAS"]
# ────────────────────────────────────────────────────────────────

try:
    from openpyxl import load_workbook
    import requests
except ImportError:
    raise SystemExit("Faltam dependências.  Rode:  pip install openpyxl requests")


# ═══════════════════════════ helpers ════════════════════════════

def pc(v):
    """Converte valor de célula para float de coordenada, ou None se inválido/zero."""
    if v is None: return None
    if isinstance(v, str):
        v = v.strip()
        if v.upper() == 'NULL' or v == '': return None
        v = v.replace(',', '.')
        try: v = float(v)
        except: return None
    if isinstance(v, (int, float)):
        return None if v == 0 else float(v)
    return None

def vsp(lat, lng):
    """Confirma que o par lat/lng está dentro do estado de SP."""
    return (lat is not None and lng is not None
            and SP_BOUNDS[0] <= lat <= SP_BOUNDS[1]
            and SP_BOUNDS[2] <= lng <= SP_BOUNDS[3])

def nz(s):
    """String → upper sem espaços; None/NULL/'' → None."""
    if s is None: return None
    s = str(s).strip()
    return None if (s.upper() == 'NULL' or s == '') else s.upper()

def strip_ac(s):
    """Remove acentos."""
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')

ABBR = {
    'JD':'JARDIM','JDM':'JARDIM','JARD':'JARDIM','VL':'VILA','V':'VILA',
    'PQ':'PARQUE','PRQ':'PARQUE','PARQ':'PARQUE','CJ':'CONJUNTO','CJTO':'CONJUNTO',
    'CONJ':'CONJUNTO','RES':'RESIDENCIAL','RESID':'RESIDENCIAL','PRES':'PRESIDENTE',
    'PROF':'PROFESSOR','PROFA':'PROFESSORA','DR':'DOUTOR','DRA':'DOUTORA',
    'STO':'SANTO','STA':'SANTA','S':'SAO','CID':'CIDADE','NSA':'NOSSA',
    'SRA':'SENHORA','SR':'SENHOR','PE':'PADRE','GAL':'GENERAL','CEL':'CORONEL',
    'CAP':'CAPITAO','ENG':'ENGENHEIRO','ALM':'ALMIRANTE','VISC':'VISCONDE',
    'MAL':'MARECHAL','AV':'AVENIDA','PCA':'PRACA','NS':'NOSSA',
}

def norm_bairro(s):
    """Normaliza nome de bairro: remove acentos e expande abreviações comuns."""
    if not s: return ''
    s = strip_ac(s.upper())
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    out = [ABBR.get(t, t) for t in s.split() if t]
    return re.sub(r'\s+', ' ', ' '.join(out)).strip()

def expand_city(s):
    """Expande nome de cidade para query de busca."""
    if not s: return ''
    s = strip_ac(s.upper()).strip()
    if s.startswith('S.') or s.startswith('S '):
        s = re.sub(r'^S\.?\s*', 'SAO ', s)
    s = s.replace('STO.', 'SANTO ').replace('STA.', 'SANTA ')
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip().title()

def expand_bairro_label(s):
    """Expande abreviações para query de busca."""
    return re.sub(r'\s+', ' ', ' '.join(
        ABBR.get(t, t)
        for t in strip_ac(str(s).upper()).replace('.', ' ').split()
    )).strip().title()


# ═══════════════════════════ Passo 1 — Leitura e agregação ══════

print("=" * 62)
print("PASSO 1 — Lendo planilha e agregando ocorrências por bairro+cidade")
print("=" * 62)

wb = load_workbook(XLSX, read_only=True)

agg   = defaultdict(lambda: {'vd':0,'fem':0,'tent':0,'slat':0.0,'slng':0.0,'n':0})
citys = defaultdict(lambda: {'slat':0.0,'slng':0.0,'n':0})
dates = []

def add(cidade, bairro, cat, lat, lng, d):
    cidade = cidade or 'NÃO INFORMADO'
    bairro = bairro or 'NÃO INFORMADO'
    a = agg[(cidade, bairro)]
    a[cat] += 1
    if vsp(lat, lng):
        a['slat'] += lat; a['slng'] += lng; a['n'] += 1
        c = citys[cidade]; c['slat'] += lat; c['slng'] += lng; c['n'] += 1
    if d is not None:
        dates.append(d)

# Aba 1 — feminicídios consumados e tentativas
ws   = wb['Base de Dados (1)']
rows = ws.iter_rows(values_only=True)
h    = list(next(rows))
ix   = {k: i for i, k in enumerate(h)}
for r in rows:
    add(nz(r[ix['CIDADE']]), nz(r[ix['BAIRRO']]),
        'fem' if r[ix['FLAG_STATUS_CRIME']] == 'C' else 'tent',
        pc(r[ix['LATITUDE']]), pc(r[ix['LONGITUDE']]),
        r[ix['DATA_OCORRENCIA_BO']])

# Aba 2 — violência doméstica geral
# (aceita tanto 'CIDADE' quanto 'NOME_MUNICIPIO', dependendo da versão da planilha)
ws   = wb['Base de Dados (2)']
rows = ws.iter_rows(values_only=True)
h    = list(next(rows))
ix   = {k: i for i, k in enumerate(h)}
cidade_col = 'CIDADE' if 'CIDADE' in ix else 'NOME_MUNICIPIO'
for r in rows:
    add(nz(r[ix[cidade_col]]), nz(r[ix['BAIRRO']]),
        'vd',
        pc(r[ix['LATITUDE']]), pc(r[ix['LONGITUDE']]),
        r[ix['DATA_OCORRENCIA_BO']])

# ── Separa os que têm coord dos que não têm ──────────────────────
bairros = []
geo_idx = defaultdict(dict)   # cidade → {norm_bairro → índice em bairros}
nogeo   = []                  # (cidade, bairro, vd, fem, tent)

for (cidade, bairro), a in agg.items():
    if a['n'] > 0:
        pos = len(bairros)
        bairros.append({
            'c': cidade, 'b': bairro,
            'lat': round(a['slat'] / a['n'], 4),
            'lng': round(a['slng'] / a['n'], 4),
            'vd': a['vd'], 'fe': a['fem'], 'te': a['tent'],
        })
        geo_idx[cidade].setdefault(norm_bairro(bairro), pos)
    else:
        nogeo.append((cidade, bairro, a['vd'], a['fem'], a['tent']))

# Tier B — mesmo par bairro+cidade, nome normalizado idêntico
residual = []
for (cidade, bairro, vd, fe, te) in nogeo:
    nb  = norm_bairro(bairro)
    pos = geo_idx.get(cidade, {}).get(nb)
    if pos is not None and nb not in ('NAO INFORMADO', ''):
        b = bairros[pos]; b['vd'] += vd; b['fe'] += fe; b['te'] += te
    else:
        residual.append((cidade, bairro, vd, fe, te))

TV = sum(a['vd']   for a in agg.values())
TF = sum(a['fem']  for a in agg.values())
TT = sum(a['tent'] for a in agg.values())
TOT = TV + TF + TT

print(f"   Pares bairro+cidade únicos: {len(agg)}")
print(f"   Com coordenada própria:     {len(bairros)}")
print(f"   Residual (sem coord):       {len(residual)}")


# ═══════════════════════════ Passo 2 — Match por prefixo (opt-in) ═

if FUZZY_PREFIX > 0:
    print()
    print("=" * 62)
    print(f"PASSO 2 — Match por prefixo normalizado ({FUZZY_PREFIX} chars)")
    print("=" * 62)

    prefix_idx = defaultdict(dict)
    for cidade, nb_map in geo_idx.items():
        for nb, pos in nb_map.items():
            if len(nb) >= FUZZY_PREFIX:
                prefix_idx[cidade].setdefault(nb[:FUZZY_PREFIX], pos)

    still_missing = []
    fuzzy_ok = 0
    for (cidade, bairro, vd, fe, te) in residual:
        nb = norm_bairro(bairro)
        if not cidade or len(nb) < FUZZY_PREFIX or nb in ('NAO INFORMADO', ''):
            still_missing.append((cidade, bairro, vd, fe, te))
            continue
        pos = prefix_idx.get(cidade, {}).get(nb[:FUZZY_PREFIX])
        if pos is not None:
            b = bairros[pos]; b['vd'] += vd; b['fe'] += fe; b['te'] += te
            fuzzy_ok += 1
        else:
            still_missing.append((cidade, bairro, vd, fe, te))

    print(f"   Associados por prefixo: {fuzzy_ok}")
    print(f"   Residual restante:      {len(still_missing)}")
    residual = still_missing
else:
    print()
    print("   (Passo 2 — match por prefixo desativado; FUZZY_PREFIX=0)")


if SOMENTE_CIDADES:
    sel = set(SOMENTE_CIDADES)
    residual = [r for r in residual if r[0] in sel]
    print(f"   Filtro de cidades aplicado: {len(residual)} registros restantes")


# ═══════════════════════════ Passo 3 — Geocodificação online ════

print()
print("=" * 62)
print("PASSO 3 — Geocodificação online via Nominatim  (Ctrl+C para pausar)")
print("=" * 62)

cache: dict = {}
if os.path.exists(CACHE):
    try:
        cache = json.load(open(CACHE, encoding='utf-8'))
        print(f"   Cache carregado: {len(cache)} entradas")
    except Exception:
        pass

session = requests.Session()
session.headers.update({'User-Agent': USER_AGENT, 'Accept-Language': 'pt-BR'})

def ck(bairro, cidade):
    return f"{norm_bairro(bairro)}|{strip_ac(str(cidade).upper())}"

def geocode(bairro, cidade):
    key = ck(bairro, cidade)
    if key in cache:
        return cache[key]
    q = f"{expand_bairro_label(bairro)}, {expand_city(cidade)}, São Paulo, Brasil"
    params = {
        'q': q, 'format': 'jsonv2', 'limit': 1, 'countrycodes': 'br',
        'viewbox': f"{SP_BOUNDS[2]},{SP_BOUNDS[1]},{SP_BOUNDS[3]},{SP_BOUNDS[0]}",
        'bounded': 1,
    }
    res = None
    for attempt in range(4):
        try:
            time.sleep(SLEEP_SEC)
            rp = session.get(GEOCODER_URL, params=params, timeout=30)
            if rp.status_code in (429, 502, 503, 504):
                time.sleep(5 * (attempt + 1)); continue
            rp.raise_for_status()
            arr = rp.json()
            if arr:
                la, lo = float(arr[0]['lat']), float(arr[0]['lon'])
                if vsp(la, lo):
                    res = [round(la, 4), round(lo, 4)]
            break
        except Exception:
            time.sleep(3 * (attempt + 1))
    cache[key] = res
    return res

# Agrupa em pares únicos para não repetir buscas
uniq: dict = {}
for (cidade, bairro, vd, fe, te) in residual:
    k = (cidade, bairro)
    if k not in uniq:
        uniq[k] = [0, 0, 0]
    uniq[k][0] += vd; uniq[k][1] += fe; uniq[k][2] += te

ja_cache = sum(1 for (c, b) in uniq if ck(b, c) in cache)
novas    = len(uniq) - ja_cache
print(f"   Pares únicos a resolver:       {len(uniq)}")
print(f"   Já em cache (instantâneo):     {ja_cache}")
print(f"   Novas buscas necessárias:      {novas}")
if novas > 0:
    print(f"   Tempo estimado para buscas:    ~{novas * SLEEP_SEC / 60:.0f} min")

ok = miss = novos_cnt = 0
total = len(uniq)
geocoded: dict = {}

for n, ((cidade, bairro), (vd, fe, te)) in enumerate(uniq.items(), 1):
    ja = ck(bairro, cidade) in cache
    coord = geocode(bairro, cidade)
    if not ja:
        novos_cnt += 1
    if coord:
        geocoded[(cidade, bairro)] = coord; ok += 1
    else:
        miss += 1
    if novos_cnt and novos_cnt % SAVE_EVERY == 0:
        json.dump(cache, open(CACHE, 'w', encoding='utf-8'), ensure_ascii=False)
    if n % 50 == 0 or n == total:
        print(f"   {n}/{total}  ok={ok}  sem_resultado={miss}  novas_buscas={novos_cnt}")

json.dump(cache, open(CACHE, 'w', encoding='utf-8'), ensure_ascii=False)
print(f"   Concluído: {ok} encontrados / {miss} sem resultado.")


# ═══════════════════════════ Passo 4 — Montar estrutura final ════

print()
print("=" * 62)
print("PASSO 4 — Montando estrutura de dados do mapa")
print("=" * 62)

# Adiciona os bairros geocodificados online
for (cidade, bairro), (vd, fe, te) in uniq.items():
    if (cidade, bairro) not in geocoded:
        continue
    lat, lng = geocoded[(cidade, bairro)]
    nb  = norm_bairro(bairro)
    pos = geo_idx.get(cidade, {}).get(nb)
    if pos is not None:
        b = bairros[pos]; b['vd'] += vd; b['fe'] += fe; b['te'] += te
    else:
        bairros.append({
            'c': cidade, 'b': bairro,
            'lat': lat, 'lng': lng,
            'vd': vd, 'fe': fe, 'te': te,
        })

# Residual final → centróide do município (fallback)
ng   = defaultdict(lambda: {'vd':0,'fe':0,'te':0,'nb':0})
unpl = {'vd':0,'fe':0,'te':0}

for (cidade, bairro), (vd, fe, te) in uniq.items():
    if (cidade, bairro) in geocoded:
        continue
    nb  = norm_bairro(bairro)
    if geo_idx.get(cidade, {}).get(nb) is not None:
        continue   # já resolvido pelo Tier B / prefixo
    c = citys.get(cidade)
    if c and c['n'] > 0:
        g = ng[cidade]
        g['vd'] += vd; g['fe'] += fe; g['te'] += te; g['nb'] += 1
    else:
        unpl['vd'] += vd; unpl['fe'] += fe; unpl['te'] += te

cidades_ng = [
    {
        'c': cidade,
        'lat': round(citys[cidade]['slat'] / citys[cidade]['n'], 4),
        'lng': round(citys[cidade]['slng'] / citys[cidade]['n'], 4),
        'vd': g['vd'], 'fe': g['fe'], 'te': g['te'], 'nb': g['nb'],
    }
    for cidade, g in ng.items()
]

# top 15 por categoria
def topn(key, k=15):
    return sorted([b for b in bairros if b[key] > 0],
                  key=lambda b: b[key], reverse=True)[:k]

top = {kk: [{'b': b['b'], 'c': b['c'], 'v': b[kk]} for b in topn(kk)]
       for kk in ('vd', 'fe', 'te')}

# metadados
ds   = [d for d in dates if d]
dmin = min(ds).strftime('%d/%m/%Y')
dmax = max(ds).strftime('%d/%m/%Y')
ncid    = len(set(b['c'] for b in bairros) | set(c['c'] for c in cidades_ng))
geo_occ = sum(b['vd'] + b['fe'] + b['te'] for b in bairros)
ng_occ  = sum(c['vd'] + c['fe'] + c['te'] for c in cidades_ng)

meta = {
    'total_vd': TV, 'total_fe': TF, 'total_te': TT,
    'n_bairros_geo': len(bairros),
    'n_cidades_ng':  len(cidades_ng),
    'n_municipios':  ncid,
    'date_min': dmin, 'date_max': dmax,
    'pct_geo':  round(100 * geo_occ / TOT, 1),
    'pct_mapa': round(100 * (geo_occ + ng_occ) / TOT, 1),
}
data = {'meta': meta, 'top': top, 'bairros': bairros, 'cidades_ng': cidades_ng}

print(f"   Bairros no mapa:                {len(bairros)}")
print(f"   Municípios no centróide:        {len(cidades_ng)}")
print(f"   Cobertura nível de bairro:      {meta['pct_geo']}%")
print(f"   Total no mapa (bairro+centróide): {meta['pct_mapa']}%")


# ═══════════════════════════ Passo 5 — Injetar no HTML ══════════

print()
print("=" * 62)
print("PASSO 5 — Atualizando o HTML")
print("=" * 62)

html = open(HTML_IN, encoding='utf-8').read()
i = html.find('window.__DADOS__ = ')
j = html.find('</script>', i)
if i < 0 or j < 0:
    raise SystemExit("Bloco de dados não encontrado no HTML. Verifique HTML_IN.")

html = (html[:i]
        + 'window.__DADOS__ = '
        + json.dumps(data, ensure_ascii=False, separators=(',', ':'))
        + ';\n'
        + html[j:])

open(HTML_OUT, 'w', encoding='utf-8').write(html)
print(f"   Salvo: {HTML_OUT}")
print()
print("✔  Concluído!")
